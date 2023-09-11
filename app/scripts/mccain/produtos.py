# flake8: noqa
import os
import logging
import openpyxl
import pandas as pd
from ftplib import FTP
from datetime import datetime
from dotenv import load_dotenv
from config import get_db_engine, log_data

load_dotenv()
log_data()


def produtos():
    FTP_CONFIG = {
        'server_ftp_mccain': os.getenv('SERVER_FTP_MCCAIN'),
        'user_ftp_mccain': os.getenv('USER_FTP_MCCAIN'),
        'password_ftp_mccain': os.getenv('PASSWORD_FTP_MCCAIN'),
        'path_clientes_mccain': os.getenv('PATH_CLIENTES_MCCAIN'),
        'path_estoque_mccain': os.getenv('PATH_ESTOQUE_MCCAIN'),
        'path_produto_mccain': os.getenv('PATH_PRODUTO_MCCAIN'),
        'path_vendas_mccain': os.getenv('PATH_VENDAS_MCCAIN'),
    }

    unid_codigos = ['001', '002', '003']

    conn = get_db_engine()
    ftp = FTP_CONFIG

    for unid_codigo in unid_codigos:
        query = (f"""
            (
                SELECT 
                    prun.prun_unid_codigo AS unidade,
                    prun.prun_ativo as tipo,
                    prun.prun_prod_codigo AS prod_codigo,
                    prod.prod_codbarras AS cod_barras,
                    prod.prod_pesoliq AS prod_pesoliq,
                    prun.prun_emb AS embalagem,
                    prod.prod_forn_codigo AS cod_fornecedor,
                    TO_CHAR(prod.prod_codigo, '00999') AS cod_prod,
                    prod.prod_descricao AS produto
                FROM produn AS prun 
                LEFT JOIN produtos AS prod ON prun.prun_prod_codigo = prod.prod_codigo
                WHERE prun.prun_bloqueado = 'N' 
                AND prun.prun_unid_codigo = '{unid_codigo}'
                AND prun.prun_ativo = 'S'
                AND prod.prod_marca IN ('MCCAIN','MCCAIN RETAIL')
            )  
        """)

        df = pd.read_sql_query(query, conn)

        wb = openpyxl.Workbook()
        ws = wb.active

        ws['A1'] = ('Code')
        ws['B1'] = ('Nome/Descrição')
        ws['C1'] = ('Quantidade x Peso')
        ws['D1'] = ('Embalagem')
        ws['E1'] = ('EAN')
        ws['F1'] = ('Cod do Fabricante')
        for index, row in df.iterrows():
            codProduto = row["cod_prod"].zfill(5)
            nomeProduto = row["produto"]
            pesoEmb = row["prod_pesoliq"]
            embalagem = row["embalagem"]
            if embalagem == 'CX':
                embalagem = 'CAIXA'
            elif embalagem == 'PC':
                embalagem = 'PACOTE'
            elif embalagem == 'UN':
                embalagem = 'UNIDADE'
            elif embalagem == 'KG':
                embalagem = 'KILOGRAMAS'
            else:
                embalagem = 'OUTROS'
            codBarras = row["cod_barras"].zfill(13)
            codFornecedor = row["cod_fornecedor"]

            ws.cell(row=index+2, column=1).value = (f'{codProduto}')
            ws.cell(row=index+2, column=2).value = (f'{nomeProduto}')
            ws.cell(row=index+2, column=3).value = (f'{pesoEmb:.2f}')
            ws.cell(row=index+2, column=4).value = (f'{embalagem}')
            ws.cell(row=index+2, column=5).value = (f'{codBarras}')
            ws.cell(row=index+2, column=6).value = (f'{codFornecedor:.0f}')

        dataAtual = datetime.now().strftime("%Y-%m-%d")
        nomeArquivo = (f'PRODUTOSDUSNEI{unid_codigo}{dataAtual}')
        ws.title = dataAtual
        diretorio = f'Z:/repositório/Dusnei-Automation/data_send/mccain/{dataAtual}'
        if not os.path.exists(diretorio):
            os.mkdir(diretorio, exist_ok=True)
        local_arquivo = os.path.join(
            f'Z:/repositório/Dusnei-Automation/data_send/mccain/{dataAtual}/{nomeArquivo}.xlsx')

        wb.save(local_arquivo)


    with FTP(FTP_CONFIG['server_ftp_mccain']) as ftp:
        ftp.login(user=FTP_CONFIG['user_ftp_mccain'], passwd=FTP_CONFIG['password_ftp_mccain'])

        remote_dir_path = os.path.join(FTP_CONFIG['path_produto_mccain'])

        # try:
        #     ftp.mkd(remote_dir_path)
        #     print(f'Diretório {remote_dir_path} criado!')
        # except Exception as e:
        #     print('Não foi possível criar a pasta, pode ser que já exista!')

        for arquivos_data in os.listdir(diretorio):
            if 'PRODUTOS' in arquivos_data:
                file_path = os.path.join(diretorio, arquivos_data)

                if os.path.isfile(file_path):
                    with open(local_arquivo, 'rb') as local_file:
                        remote_path = os.path.join(remote_dir_path, arquivos_data)
                        ftp.storbinary(f"STOR {remote_path}", local_file)
                logging.info(
                    f"Arquivo {os.path.basename(arquivos_data)} upload FTP server concluído com sucesso!")


if __name__ == "__main__":
    produtos()