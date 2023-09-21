# flake8: noqa
import os
import openpyxl
import pandas as pd
from ftplib import FTP
from datetime import datetime
from dotenv import load_dotenv
from conn import ftp_config, config
import sys
sys.path.append('Z:/repositório/Dusnei-Automation/log')
import log_config


load_dotenv()


def estoques():
    logger = log_config.setup_logger('mccain_log.log')
    FTP_CONFIG = ftp_config.FTP_CONFIG

    unid_codigos = ['001', '002', '003']

    conn = config.get_db_engine()
    ftp = FTP_CONFIG

    for unid_codigo in unid_codigos:
        query = (f"""
            (
                SELECT
                    prun.prun_estoque1 AS estoque,
                    (prun.prun_estoque1 * prod.prod_pesoliq) AS qtde,
                    prun.prun_unid_codigo AS unidade,
                    prun.prun_ativo as tipo,
                    prun.prun_prod_codigo AS prod_codigo,
                    prod.prod_codbarras AS cod_barras,
                    prod.prod_marca AS marca,
                    prod.prod_codigo AS cod_prod
                FROM produn AS prun
                LEFT JOIN produtos AS prod ON prun.prun_prod_codigo = prod.prod_codigo
                WHERE prun.prun_bloqueado = 'N'
                AND prun.prun_unid_codigo = '{unid_codigo}'
                AND prun.prun_ativo = 'S'
                AND prun.prun_estoque1 > 0
                AND prod.prod_marca IN ('MCCAIN','MCCAIN RETAIL')
            )
        """)

        df = pd.read_sql_query(query, conn)

        wb = openpyxl.Workbook()
        ws = wb.active

        dataAtualEstoque = datetime.now().strftime("%Y-%m-%d")
        ws['A1'] = ('Code')
        ws['B1'] = ('Quantity')
        ws['C1'] = ('Stock Date')
        ws['D1'] = ('Expiration Date')
        for index, row in df.iterrows():
            code = row["cod_prod"]
            quantity = row["qtde"]
            stockDate = dataAtualEstoque
            expirationDate = ''

            ws.cell(row=index+2, column=1).value = (f'{code:.0f}')
            ws.cell(row=index+2, column=2).value = (f'{quantity:.2f}')
            ws.cell(row=index+2, column=3).value = (f'{stockDate}')
            ws.cell(row=index+2, column=4).value = (f'{expirationDate}')

        dataAtual = datetime.now().strftime("%Y-%m-%d")
        nomeArquivo = (f'ESTOQUEDUSNEI{unid_codigo}{dataAtual}')
        ws.title = dataAtual
        diretorio = f'Z:/repositório/Dusnei-Automation/data_send/mccain/{dataAtual}'
        if not os.path.exists(diretorio):
            os.mkdir(diretorio, exist_ok=True)
        local_arquivo = os.path.join(
            f'Z:/repositório/Dusnei-Automation/data_send/mccain/{dataAtual}/{nomeArquivo}.xlsx')

        wb.save(local_arquivo)
    logger.info('Arquivo estoques_unidades.xlsx criado!')


    with FTP(FTP_CONFIG['server_ftp_mccain']) as ftp:
        ftp.login(user=FTP_CONFIG['user_ftp_mccain'], passwd=FTP_CONFIG['password_ftp_mccain'])
        remote_dir_path = os.path.join(FTP_CONFIG['path_estoque_mccain'])
        
        for arquivos_data in os.listdir(diretorio):
            if 'ESTOQUE' in arquivos_data:
                file_path = os.path.join(diretorio, arquivos_data)
                if os.path.isfile(file_path):
                    with open(local_arquivo, 'rb') as local_file:
                        remote_path = os.path.join(remote_dir_path, arquivos_data)
                        ftp.storbinary(f"STOR {remote_path}", local_file)
                logger.info(
                    f"Arquivo {os.path.basename(arquivos_data)} upload FTP server concluído com sucesso!")


if __name__ == "__main__":
    estoques()