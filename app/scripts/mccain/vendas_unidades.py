# flake8: noqa
import os
import openpyxl
import pandas as pd
from ftplib import FTP
from typing import List
from datetime import datetime
from dotenv import load_dotenv
from conn import ftp_config, config
import sys
sys.path.append(os.getenv('DUSNEI_LOG_DIRECTORY'))  # type: ignore


load_dotenv()


def generate_sql_query() -> List[str]:
    current_month = f"{datetime.now().month:02}"
    current_year = datetime.now().year % 100

    last_month = f"{(datetime.now().month - 1) % 12:02}"
    
    if last_month == "00":
        last_month = "12"

    last_year = current_year if datetime.now().month > 1 else current_year - 1

    tables = [f"movprodd{last_month}{last_year}", f"movprodd{current_month}{current_year}"]

    return tables



def vendas():
    FTP_CONFIG = ftp_config.FTP_CONFIG

    unid_codigos = ['001', '002', '003']

    for unid_codigo in unid_codigos:

        def vendas_query(table_name, conn, unid_codigo):
            query = (f"""
                (
                    SELECT
                        mprd.mprd_dcto_codigo AS doc_cod, 
                        mprd.mprd_transacao AS transacao,
                        clie.clie_cnpjcpf AS cnpj_cpf,
                        clie.clie_codigo AS cod_clie,
                        mprd.mprd_datamvto AS data,
                        mprd.mprd_numerodcto AS nfe,
                        prod.prod_codbarras AS cod_barras,
                        prod.prod_codigo AS cod_prod,
                        (mprd.mprd_qtde * prod.prod_pesoliq) AS quantity,
                        mprd.mprd_valor AS amount,
                        mprc.mprc_vend_codigo AS cod_vend,
                        SUBSTRING(clie.clie_cepres, 1,5) ||'-'|| SUBSTRING(clie.clie_cepres, 6,3) AS cep
                    FROM {table_name} AS mprd 
                    LEFT JOIN movprodc AS mprc ON mprd.mprd_operacao = mprc.mprc_operacao
                    LEFT JOIN produtos AS prod ON mprd.mprd_prod_codigo = prod.prod_codigo
                    LEFT JOIN clientes AS clie ON mprc.mprc_codentidade = clie.clie_codigo
                    WHERE mprd_status = 'N' 
                    AND mprd_unid_codigo IN ('{unid_codigo}')
                    AND prod.prod_marca IN ('MCCAIN','MCCAIN RETAIL')
                    AND mprd.mprd_dcto_codigo IN (
                        '6666','6667','6668','6670','7340','7335','7337','7338','7339','7341','7345',
                        '6680','6681','7321','7323','7326','7332',
                        '6890','6892','7322','7346',
                        '7260','7262','7263','7264','7268','7269'
                    )
                    AND mprd.mprd_datamvto > CURRENT_DATE - INTERVAL '10 DAYS'
                )  
            """)
                    # AND mprd.mprd_datamvto > CURRENT_DATE - INTERVAL '7 DAYS'
            return pd.read_sql_query(query, conn)

        conn = config.get_db_engine()
        ftp = FTP_CONFIG

        wb = openpyxl.Workbook()
        ws = wb.active

        tables = generate_sql_query()

        filtered_dfs = [df.dropna(how='all', axis=1) for df in [vendas_query(table, conn, unid_codigo) for table in tables]]
        df = pd.concat(filtered_dfs)

        ws['A1'] = ('systemId')
        ws['B1'] = ('Code')
        ws['C1'] = ('Quantity')
        ws['D1'] = ('Amount')
        ws['E1'] = ('Sale Date')
        ws['F1'] = ('Transaction ID')
        for index, row in df.iterrows():
            systemId = row["cod_clie"]
            code = row["cod_prod"]
            doc_cod = row["doc_cod"]
            quantity = row["quantity"]

            if doc_cod in ['7260','7262','7263','7264','7268','7269']:
                quantity = -quantity

            amount = str(row["amount"]).replace(',', '.')
            amount2 = float(amount)
            data = row["data"].strftime("%Y-%m-%d")
            transactionId = "1" + row["nfe"].zfill(6)

            ws.cell(row=index+2, column=1).value = (f'{systemId:.0f}')
            ws.cell(row=index+2, column=2).value = (f'{code:.0f}')
            ws.cell(row=index+2, column=3).value = (f'{quantity:.2f}')
            ws.cell(row=index+2, column=4).value = (f'{amount2:.2f}')
            ws.cell(row=index+2, column=5).value = (f'{data}')
            ws.cell(row=index+2, column=6).value = (f'{transactionId}')

        dataAtual = datetime.now().strftime("%Y-%m-%d")
        nomeArquivo = (f'VENDASDUSNEI{unid_codigo}{dataAtual}')
        ws.title = dataAtual
        diretorio = f'{os.getenv("DUSNEI_DATA_DIRECTORY_MCCAIN")}/{dataAtual}'
        if not os.path.exists(diretorio):
            os.mkdir(diretorio)
        local_arquivo = os.path.join(
            f'{os.getenv("DUSNEI_DATA_DIRECTORY_MCCAIN")}/{dataAtual}/{nomeArquivo}.xlsx')

        wb.save(local_arquivo)
    print('Arquivo vendas_unidades.xlsx criado!')


    with FTP(FTP_CONFIG['server_ftp_mccain']) as ftp:
        ftp.login(user=FTP_CONFIG['user_ftp_mccain'], passwd=FTP_CONFIG['password_ftp_mccain'])

        remote_dir_path = os.path.join(FTP_CONFIG['path_vendas_mccain'])

        for arquivos_data in os.listdir(diretorio):
            if 'VENDAS' in arquivos_data:
                file_path = os.path.join(diretorio, arquivos_data)

                if os.path.isfile(file_path):
                    with open(local_arquivo, 'rb') as local_file:
                        remote_path = os.path.join(remote_dir_path, arquivos_data)
                        ftp.storbinary(f"STOR {remote_path}", local_file)
                print(
                    f"Arquivo {os.path.basename(arquivos_data)} upload FTP server concluído com sucesso!")


if __name__ == "__main__":
    vendas()
  