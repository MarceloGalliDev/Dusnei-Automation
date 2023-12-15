# flake8: noqa W293
# pylint: disable=all
import os
import time
from datetime import datetime
import locale
import openpyxl
import pandas as pd
from conn import DatabaseConnection
from dotenv import load_dotenv
from loguru import logger

load_dotenv()

class PendenciasClientes():
    def __init__(self):
        load_dotenv()
        self.path_dados = os.getenv('DUSNEI_DATA_DIRECTORY_RELATORIO_PENDENCIAS')
        self.conn = DatabaseConnection.get_db_engine(self)
        self.supervisores = []
        self.vendedores = []
    
    def pendencias_clientes_query(self, conn, cod_supe):
        query = (f"""
            SELECT
                pfin.pfin_transacao AS transacao,
                pfin.pfin_status AS status,
                pfin.pfin_datavcto AS datavcto,
                pfin.pfin_pger_conta AS pger_conta,
                pfin.pfin_unid_codigo AS unid_cod,
                pfin.pfin_vend_codigo AS vend_cod,
                pfin.pfin_codentidade AS entidade_cod,
                pfin.pfin_cnpjcpf AS cnpj,
                pfin.pfin_numerodcto AS nota,
                pfin.pfin_parcela AS parcelas,
                pfin.pfin_valor AS valor,
                pfin.pfin_nparcelas AS nparcelas,
                pger.pger_descricao AS descricao_contabil,
                pger.pger_tipo AS tipo,
                clie.clie_nome AS clie_nome,
                clie.clie_razaosocial AS razaosocial,
                clie.clie_vend_codigo AS clie_vend_cod,
                clie.clie_foneres AS fone_res,
                clie.clie_fonecel AS fone_cel,
                vend.vend_nome AS vend_nome,
                vend.vend_supe_codigo AS supe_cod,
                CONCAT(vend.vend_extra16, vend.vend_extra15) AS vend_cel,
                unid.unid_reduzido AS unidade,
                supe.supe_nome AS supe_nome
            FROM pendfin AS pfin
            INNER JOIN planoger AS pger ON pfin.pfin_pger_conta = pger.pger_conta
            INNER JOIN unidades AS unid ON pfin.pfin_unid_codigo = unid.unid_codigo
            INNER JOIN clientes AS clie ON pfin.pfin_codentidade = clie.clie_codigo
            INNER JOIN vendedores AS vend ON clie.clie_vend_codigo = vend.vend_codigo
            INNER JOIN supervisores AS supe ON vend.vend_supe_codigo = supe.supe_codigo
            WHERE pfin.pfin_status = 'P'
            AND pfin.pfin_datavcto < CURRENT_DATE
            AND vend.vend_supe_codigo = {cod_supe}
            ORDER BY pfin.pfin_datavcto DESC
        """)
        df = pd.read_sql_query(query, conn)
        return df
    
    def save_to_excel_and_txt(self, cod_supe, data_atual):
        wb = openpyxl.Workbook()
        ws = wb.active
        
        data_titulo = datetime.now().strftime("%Y%m%d")
        
        locale.setlocale(locale.LC_MONETARY, 'pt-BR.UTF-8')
        
        df = self.pendencias_clientes_query(self.conn, cod_supe)
    
        grouped_df = df.groupby('supe_cod')
        
        for index, row in enumerate(group.itertuples(), start=2):
            status = "-"
            hora = "-"

            ws.cell(row=index, column=1).value = row.Index
            ws.cell(row=index, column=2).value = row.entidade_cod
            ws.cell(row=index, column=3).value = row.razaosocial
            ws.cell(row=index, column=4).value = row.cnpj
            ws.cell(row=index, column=5).value = row.nota
            ws.cell(row=index, column=6).value = row.valor
            ws.cell(row=index, column=7).value = format(row.datavcto, "%Y/%m/%d")
            ws.cell(row=index, column=8).value = row.fone_res
            ws.cell(row=index, column=9).value = row.fone_cel
            ws.cell(row=index, column=10).value = row.supe_cod
            ws.cell(row=index, column=11).value = row.supe_nome
            ws.cell(row=index, column=12).value = row.clie_vend_cod
            ws.cell(row=index, column=13).value = row.vend_nome
            ws.cell(row=index, column=14).value = row.vend_cel
            ws.cell(row=index, column=15).value = row.transacao
            ws.cell(row=index, column=16).value = int(row.pger_conta)
            ws.cell(row=index, column=17).value = row.descricao_contabil
            ws.cell(row=index, column=18).value = status
            ws.cell(row=index, column=19).value = hora
        
        if unid_codigo == ['003','010']:
            unid_codigo = '003'
        nome_arquivo = (f'PRODUTOSDUSNEI{unid_codigo}{data_atual}')
        ws.title = data_atual
        data_pasta = datetime.now().strftime("%Y-%m-%d")
        diretorio = f'{self.path_dados}/{data_pasta}'
        if not os.path.exists(diretorio):
                os.mkdir(diretorio)
        local_arquivo = os.path.join(f'{diretorio}/{nome_arquivo}.xlsx')
        wb.save(local_arquivo)
        
        time.sleep(2)
        
        local_arquivo_txt = os.path.join(f'{diretorio}/{nome_arquivo}.txt')
        with open(local_arquivo_txt, 'w') as txt_file:
            txt_file.write((f'HCADPROD   {data_titulo}') + '\n')
            txt_file.write((f'{caracter_adc}{cnpj_unid}') + '\n')
            for row in processed_rows:
                txt_file.write(row + '\n')
        
        logger.info('Arquivo produtos OK!')
    
    def produtos(self):
        for unid_codigo in self.unid_codigos:
            df = pd.concat([self.produtos_query( self.conn, unid_codigo)])
            
            processed_rows = self.process_rows(df, unid_codigo)
            data_atual = datetime.now().strftime("%Y%m%d%H%M%S%f")[:-3]
            self.save_to_excel_and_txt(processed_rows, unid_codigo, data_atual)
        
        logger.info('Funções produtos OK!')
            
if __name__ == "__main__":
    db_clientes = Produtos()    
    db_clientes.produtos()