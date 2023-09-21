# flake8: noqa
import os
import openpyxl
import pandas as pd
from ftplib import FTP
from datetime import datetime
from dotenv import load_dotenv
from conn import ftp_config, config
import sys
sys.path.append('Z:/repositório/Dusnei-Automation/log')
import log_config


load_dotenv()


def vendas_estado():
    logger = log_config.setup_logger('mccain_log.log')
    FTP_CONFIG = ftp_config.FTP_CONFIG

    cod_estados = ['PR','SP']

    for cod_estado in cod_estados:

        def vendas_query(table_name, conn, cod_estado):
            query = (f"""
                (
                    SELECT
                        mprd.mprd_dcto_codigo AS doc_cod, 
                        mprd.mprd_transacao AS transacao,
                        clie.clie_cnpjcpf AS cnpj_cpf,
                        clie.clie_codigo AS cod_clie,
                        mprd.mprd_datamvto AS data,
                        mprd.mprd_numerodcto AS nfe,
                        prod.prod_codbarras AS cod_barras,
                        prod.prod_codigo AS cod_prod,
                        (mprd.mprd_qtde * prod.prod_pesoliq) AS quantity,
                        mprd.mprd_valor AS amount,
                        mprc.mprc_vend_codigo AS cod_vend,
                        mprc.mprc_uf AS estado,
                        prod.prod_marca AS marca,
                        SUBSTRING(clie.clie_cepres, 1,5) ||'-'|| SUBSTRING(clie.clie_cepres, 6,3) AS cep
                    FROM {table_name} AS mprd 
                    LEFT JOIN movprodc AS mprc ON mprd.mprd_operacao = mprc.mprc_operacao
                    LEFT JOIN produtos AS prod ON mprd.mprd_prod_codigo = prod.prod_codigo
                    LEFT JOIN clientes AS clie ON mprc.mprc_codentidade = clie.clie_codigo
                    WHERE mprd_status = 'N' 
                    AND prod.prod_marca IN ('MCCAIN','MCCAIN RETAIL')
                    AND mprd.mprd_dcto_codigo IN ('6666','6668','7335','7337','7338','7339','7260','7263','7262','7268','7264','7269', '7267', '7319', '7318')
                    AND mprc.mprc_uf = '{cod_estado}'
                    AND mprd.mprd_datamvto >= '2023-09-01'
                    AND mprd.mprd_datamvto <= '2023-09-20'
                )  
            """)
                    # AND mprd.mprd_datamvto > CURRENT_DATE - INTERVAL '8 DAYS'
            return pd.read_sql_query(query, conn)


        conn = config.get_db_engine()
        ftp = FTP_CONFIG

        wb = openpyxl.Workbook()
        ws = wb.active

        tables = [
            'movprodd0123', 
            'movprodd0223', 
            'movprodd0323', 
            'movprodd0423', 
            'movprodd0523', 
            'movprodd0623', 
            'movprodd0723', 
            'movprodd0823', 
            'movprodd0923', 
            'movprodd1023', 
            'movprodd1123', 
            'movprodd1223'
        ]

        df = pd.concat([vendas_query(table, conn, cod_estado)for table in tables])

        ws['A1'] = ('systemId')
        ws['B1'] = ('Code')
        ws['C1'] = ('Quantity')
        ws['D1'] = ('Amount')
        ws['E1'] = ('Sale Date')
        ws['F1'] = ('Transaction ID')
        for index, row in df.iterrows():
            systemId = row["cod_clie"]
            code = row["cod_prod"]
            doc_cod = row["doc_cod"]
            quantity = row["quantity"]

            if doc_cod in ['7260', '7263', '7262', '7268', '7264', '7269', '7267', '7319', '7318']:
                quantity = -quantity

            amount = str(row["amount"]).replace(',', '.')
            amount2 = float(amount)
            data = row["data"].strftime("%Y-%m-%d")
            transactionId = "1" + row["nfe"].zfill(6)

            ws.cell(row=index+2, column=1).value = (f'{systemId:.0f}')
            ws.cell(row=index+2, column=2).value = (f'{code:.0f}')
            ws.cell(row=index+2, column=3).value = (f'{quantity:.2f}')
            ws.cell(row=index+2, column=4).value = (f'{amount2:.2f}')
            ws.cell(row=index+2, column=5).value = (f'{data}')
            ws.cell(row=index+2, column=6).value = (f'{transactionId}')

        dataAtual = datetime.now().strftime("%Y-%m-%d")
        nomeArquivo = (f'VENDASDUSNEI{cod_estado}{dataAtual}')
        ws.title = dataAtual
        diretorio = f'Z:/repositório/Dusnei-Automation/data_send/mccain/{dataAtual}'
        if not os.path.exists(diretorio):
            os.mkdir(diretorio)
        local_arquivo = os.path.join(
            f'Z:/repositório/Dusnei-Automation/data_send/mccain/{dataAtual}/{nomeArquivo}.xlsx')

        wb.save(local_arquivo)
        logger.info('Arquivo vendas_estados.xlsx criado!')


    with FTP(FTP_CONFIG['server_ftp_mccain']) as ftp:
        ftp.login(user=FTP_CONFIG['user_ftp_mccain'], passwd=FTP_CONFIG['password_ftp_mccain'])

        remote_dir_path_pr = os.path.join(FTP_CONFIG['path_vendas_pr_mccain'])
        remote_dir_path_sp = os.path.join(FTP_CONFIG['path_vendas_sp_mccain'])

        # try:
        #     ftp.mkd(remote_dir_path)
        #     print(f'Diretório {remote_dir_path} criado!')
        # except Exception as e:
        #     print('Não foi possível criar a pasta, pode ser que já exista!')

        for arquivos_data in os.listdir(diretorio):
            if 'VENDASDUSNEIPR' in arquivos_data:
                file_path = os.path.join(diretorio, arquivos_data)

                if os.path.isfile(file_path):
                    with open(local_arquivo, 'rb') as local_file:
                        remote_path = os.path.join(remote_dir_path_pr, arquivos_data)
                        ftp.storbinary(f"STOR {remote_path}", local_file)
                logger.info(
                    f"Arquivo {os.path.basename(arquivos_data)} upload FTP server concluído com sucesso!")
                
            if 'VENDASDUSNEISP' in arquivos_data:
                file_path = os.path.join(diretorio, arquivos_data)

                if os.path.isfile(file_path):
                    with open(local_arquivo, 'rb') as local_file:
                        remote_path = os.path.join(remote_dir_path_sp, arquivos_data)
                        ftp.storbinary(f"STOR {remote_path}", local_file)
                logger.info(
                    f"Arquivo {os.path.basename(arquivos_data)} upload FTP server concluído com sucesso!")


if __name__ == "__main__":
    vendas_estado()
  