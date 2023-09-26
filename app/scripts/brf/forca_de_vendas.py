# flake8: noqa W293
# pylint: disable=all
import os
import time
import openpyxl
import pandas as pd
from loguru import logger
from datetime import datetime
from dotenv import load_dotenv
from conn import DatabaseConnection
from unidecode import unidecode

load_dotenv()

class Forca_De_Venda:
    def __init__(self):
        load_dotenv()
        self.path_dados = os.getenv('DUSNEI_DATA_DIRECTORY_BRF')
        self.unid_codigos = ["001", "002", ['003','010']]
        self.conn = DatabaseConnection.get_db_engine(self)
    
    def forca_de_vendas_query(self, conn, unid_codigo):
        if isinstance(unid_codigo, list):
            unid_values = ",".join([f"'{code}'" for code in unid_codigo])
        else:
            unid_values = f"'{unid_codigo}'"
        query = (f"""
            SELECT 
                clie.clie_unid_codigo AS unidade,
                clie.clie_codigo AS clie_codigo,
                clie.clie_nome AS clie_nome,
                clie.clie_cnpjcpf AS cnpj_cpf,
                vend.vend_supe_codigo AS supe_codigo,
                LPAD(clie.clie_vend_codigo, 4, '0') AS cod_vendedor,
                REPLACE(UPPER(vend.vend_nome),' ','') AS nome_vendedor,
                TO_CHAR(vend.vend_codfrente, '0000') AS cod_gerente,
                REPLACE(UPPER(vend.vend_extra8),' ','') AS nome_gerente,
                TO_CHAR(supe.supe_codigo::numeric, '0000') AS cod_supervisor,
                REPLACE(UPPER(supe.supe_nome),' ','') AS nome_supervisor
            FROM clientes AS clie 
            LEFT JOIN vendedores AS vend ON clie.clie_vend_codigo = vend.vend_codigo
            LEFT JOIN supervisores AS supe ON vend.vend_supe_codigo = supe.supe_codigo
            LEFT JOIN municipios AS muni ON clie.clie_muni_codigo_res = muni.muni_codigo
            WHERE clie.clie_tipos NOT IN ('','VE','FU','UN','NL')
            AND clie.clie_endres NOT IN ('') 
            AND muni.muni_nome NOT IN ('','IDENTIFICAR', 'Identificar') 
            AND clie.clie_rota_codigo NOT IN ('') 
            AND clie.clie_unid_codigo IN ({unid_values})
            AND clie.clie_cnpjcpf > '0'
            AND clie.clie_cnpjcpf <> ''
            AND clie.clie_cepres NOT IN ('00000-000','','0','00000','00000000')
            AND clie.clie_cepres NOT IN ('')
        """)
        return pd.read_sql_query(query, conn)

    def process_rows(self, df, unid_codigo):
        processed_rows = []
        
        def sanitize_string_name(value, default_value, max_length):
            if value is None or value == '':
                return default_value.ljust(max_length)[:max_length].upper()
            return unidecode(value)[:max_length].ljust(max_length).upper()

        def sanitize_string_number(value, default_value, max_length):
            if value is None:
                return default_value.ljust(max_length)[:max_length]
            return unidecode(value)[:max_length].ljust(max_length)

        for index, row in df.iterrows():
            caracter_adc = "D"
            
            if unid_codigo == '001':
                cnpj_unid = '81894255000147'
            elif unid_codigo == '002':
                cnpj_unid = '81894255000228'
            else:
                cnpj_unid = '81894255000309'
            
            cnpj_cliente = sanitize_string_number(row["cnpj_cpf"], '0'*14, 17)
                                       
            nome_gerente_bruta = row["nome_gerente"]
            nome_supervisor_bruta = row["nome_supervisor"]
            nome_vendedor_bruta = row["nome_vendedor"]
            cod_gerente_bruta = row["cod_gerente"]
            cod_supervisor_bruta = row["cod_supervisor"]
            cod_vendedor_bruta = row["cod_vendedor"]
            
            if (
                nome_vendedor_bruta is not None or 
                nome_vendedor_bruta != '' or
                cod_vendedor_bruta is not None or
                cod_vendedor_bruta != '' or
                cod_vendedor_bruta != '000' or
                cod_vendedor_bruta != '0000'
            ):
                cod_gerente = sanitize_string_number(cod_gerente_bruta, '0000', 14)
                nome_gerente = sanitize_string_name(nome_gerente_bruta, 'INATIVADO', 49)
                cod_supervisor = sanitize_string_number(cod_supervisor_bruta, '0000', 14)
                nome_supervisor = sanitize_string_name(nome_supervisor_bruta, 'INATIVADO', 50)
                cod_vendedor = sanitize_string_number(cod_vendedor_bruta, '0000', 20)
                nome_vendedor = sanitize_string_name(nome_vendedor_bruta, 'INATIVADO', 50)
            else:
                nome_gerente = 'INATIVADO'
                cod_gerente = '0000'
                nome_supervisor = 'INATIVADO'
                cod_supervisor = '0000'
                nome_vendedor = 'INATIVADO'
                cod_vendedor = '0000'
            
            espaco_branco1 = ' '*4
            
            processed_row = (f'{caracter_adc}{cnpj_unid}{cnpj_cliente}{cod_gerente}{nome_gerente}{cod_supervisor}{nome_supervisor}{cod_vendedor}{nome_vendedor}')
            processed_rows.append(processed_row)
        logger.info('Query estoques OK!')
        logger.info('Processamento de dados estoques OK!')
        return processed_rows
    
    def save_to_excel_and_txt(self, processed_rows, unid_codigo, data_atual):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws['A1'] = ('HFV10   01838723010513')
        for index, row_value in enumerate(processed_rows, start=2):
            ws.cell(row=index, column=1).value = row_value
        
        if unid_codigo == ['003','010']:
            unid_codigo = '003'
        nome_arquivo = (f'FORCAVENDASDUSNEI{unid_codigo}{data_atual}')
        ws.title = data_atual
        data_pasta = datetime.now().strftime("%Y-%m-%d")
        diretorio = f'{self.path_dados}/{data_pasta}'
        if not os.path.exists(diretorio):
                os.mkdir(diretorio)
        local_arquivo = os.path.join(f'{diretorio}/{nome_arquivo}.xlsx')
        wb.save(local_arquivo)
        
        time.sleep(2)
        
        local_arquivo_txt = os.path.join(f'{diretorio}/{nome_arquivo}.txt')
        with open(local_arquivo_txt, 'w') as txt_file:
            txt_file.write('HFV10   01838723010513' + '\n')
            for row in processed_rows:
                txt_file.write(row + '\n')
        
        logger.info('Arquivo forca_de_vendas OK!')
    
    def forca_de_vendas(self):
        for unid_codigo in self.unid_codigos:
            df = pd.concat([self.forca_de_vendas_query( self.conn, unid_codigo)])
            
            processed_rows = self.process_rows(df, unid_codigo)
            data_atual = datetime.now().strftime("%Y%m%d%H%M%S%f")[:-3]
            self.save_to_excel_and_txt(processed_rows, unid_codigo, data_atual)
        
        logger.info('Funções forca_de_vendas OK!')
            
if __name__ == "__main__":
    db_clientes = Forca_De_Venda()    
    db_clientes.forca_de_vendas()